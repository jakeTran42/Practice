import openpyxl, os

'''
Create and Modify Excel Spread Sheets
'''

wb = openpyxl.Workbook() #Creating a new excel spreadhseet, workbook object will be returned

wb.get_sheet_names() # New Workbook will have one sheet to work with

sheet = wb.get_sheet_by_name('Sheet')
sheet['A1'] = 42 # Editing the cell of the sheet
sheet['A2'] = 'Hello'

os.getcwd()
os.chdir(r'C:\Users\Udemy\AutomatingBoringPython\Excel-Word-PDF')
wb.save('example1.xlsx') #Save the new workbook to current dir


sheet2 = wb.create_sheet() #create worksheet and return that sheet object

wb.get_sheet_names()

sheet2.title = 'MyNewSheet' #change sheet name

wb.save('example1.xlsx')

wb.create_sheet(index=0, title='SheetNo1')

wb.save('example1.xlsx')
