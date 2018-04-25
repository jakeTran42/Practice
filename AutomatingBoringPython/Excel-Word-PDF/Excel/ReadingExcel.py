import openpyxl
import os

'''
Reading an Excel Workbook
'''

os.getcwd()
os.chdir(r'C:\Users\Insui\Desktop\dev-stuff\Udemy\AutomatingBoringPython\Excel-Word-PDF')
os.getcwd()

workbook = openpyxl.load_workbook('example.xlsx') #Opening the example workbook

type(workbook) #See what type 'workbook' is. it is a workbook object

workbook.get_sheet_names() # list out all sheet

sheet = workbook.get_sheet_by_name('Sheet1')

cell = sheet['C1'] # return a cell object for us to work with. We're getting the cell A1 in this case

cell.value #The cell well have a member variable value

str(cell.value) #Getting onlu str value for the cell

sheet.cell(row=1, column=2) #Specifying what row and column we want with cell method

for i in range(1, 8):
    print('row %s, value: %s' % (i, sheet.cell(row=i, column=2).value))
